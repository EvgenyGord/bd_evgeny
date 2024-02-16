from django.shortcuts import render, get_object_or_404, redirect
from .forms import AuthorForm, EmployeeForm, GenreForm, SubscriberForm, BookForm, BookDistributionForm
from .models import Author, Employee, Genre, Subscriber, Book, BookDistribution
from django.http import HttpResponse
import openpyxl
from django.db import connection
from django.contrib.auth.decorators import login_required

# Create your views here.

def home(request):
    # Извлекаем значение параметра sql_query из запроса
    sql_query = request.GET.get('sql_query', '')
    # Передаем значение параметра в контекст
    return render(request, 'biblioteka/home.html', {'sql_query': sql_query})


@login_required
def display_authors(request):
    authors = Author.objects.all().order_by('author_id')
    return render(request, 'authors/authors_list.html', {'authors': authors})
@login_required
def add_author(request):
    if request.method == 'POST':
        form = AuthorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_authors')
    else:
        form = AuthorForm()

    return render(request, 'authors/add_author.html', {'form': form})
@login_required
def edit_author(request, author_id):
    author = get_object_or_404(Author, author_id=author_id)

    if request.method == 'POST':
        form = AuthorForm(request.POST, instance=author)
        if form.is_valid():
            form.save()
            return redirect('display_authors')
    else:
        form = AuthorForm(instance=author)

    return render(request, 'authors/edit_author.html', {'form': form, 'author': author})
@login_required
def delete_author(request, author_id):
    author = get_object_or_404(Author, author_id=author_id)
    author.delete()
    return redirect('display_authors')
@login_required
def export_authors_to_excel(request):
    authors = Author.objects.all().order_by('author_id')
    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    # Добавляем заголовки
    ws.append(['id', 'Имя', 'Фамилия', 'Год рождения', 'Год смерти'])
    # Добавляем данные
    for author in authors:
        ws.append([author.author_id, author.first_name, author.last_name, author.year_of_birth, author.year_of_death])
    # Сохраняем книгу и отправляем пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=authors.xlsx'
    wb.save(response)
    return response
@login_required
def report_authors(request):
    authors = Author.objects.all().order_by('author_id')
    return render(request, 'authors/report_authors.html', {'authors': authors})
@login_required
def execute_sql_query(request):
    if request.method == 'POST':
        sql_query = request.POST.get('sql_query', '')
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
        except Exception as e:
            return render(request, 'authors/sql_result.html', {'sql_error': str(e)})

        # Добавим данные в контекст для передачи в шаблон
        context = {'sql_columns': columns, 'sql_rows': rows, 'sql_query': sql_query}
        return render(request, 'authors/sql_result.html', context)
    return render(request, 'biblioteka/sql_result.html')
@login_required
def export_sql_results_to_excel(request):
    sql_query = request.GET.get('query', '')

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
    except Exception as e:
        return HttpResponse(f'Ошибка выполнения SQL-запроса: {str(e)}')

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Добавляем заголовки
    ws.append(columns)

    # Добавляем данные
    for row in rows:
        ws.append(row)

    # Сохраняем книгу и отправляем пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=sql_results.xlsx'
    wb.save(response)

    return response


#Employee------------------------------------------------------------------------------------------------------------------------------------------------


@login_required
def display_employees(request):
    employees = Employee.objects.all()
    return render(request, 'employees/employees_list.html', {'employees': employees})
@login_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_employees')
    else:
        form = EmployeeForm()

    return render(request, 'employees/add_employee.html', {'form': form})
@login_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, employee_id=employee_id)

    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('display_employees')
    else:
        form = EmployeeForm(instance=employee)

    return render(request, 'employees/edit_employee.html', {'form': form, 'employee': employee})
@login_required
def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, employee_id=employee_id)
    employee.delete()
    return redirect('display_employees')
@login_required
def export_employees_to_excel(request):
    employees = Employee.objects.all()

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Добавляем заголовки
    ws.append(['id', 'Имя', 'Фамилия', 'Должность', 'Номер телефона'])

    # Добавляем данные
    for employee in employees:
        ws.append([employee.employee_id, employee.first_name, employee.last_name, employee.position, employee.phone_number])

    # Сохраняем книгу и отправляем пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    wb.save(response)

    return response
@login_required
def report_employees(request):
    employees = Employee.objects.all()
    return render(request, 'employees/report_employees.html', {'employees': employees})


#Genre-----------------------------------------------------------------------------------------------------------

@login_required
def display_genres(request):
    genres = Genre.objects.all()
    return render(request, 'genres/genres_list.html', {'genres': genres})
@login_required
def add_genre(request):
    if request.method == 'POST':
        form = GenreForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_genres')
    else:
        form = GenreForm()

    return render(request, 'genres/add_genre.html', {'form': form})
@login_required
def edit_genre(request, genre_id):
    genre = get_object_or_404(Genre, genre_id=genre_id)

    if request.method == 'POST':
        form = GenreForm(request.POST, instance=genre)
        if form.is_valid():
            form.save()
            return redirect('display_genres')
    else:
        form = GenreForm(instance=genre)

    return render(request, 'genres/edit_genre.html', {'form': form, 'genre': genre})
@login_required
def delete_genre(request, genre_id):
    genre = get_object_or_404(Genre, genre_id=genre_id)
    genre.delete()
    return redirect('display_genres')
@login_required
def export_genres_to_excel(request):
    genres = Genre.objects.all()

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Добавляем заголовки
    ws.append(['id', 'Название', 'Тип', 'Рейтинг', 'Описание'])

    # Добавляем данные
    for genre in genres:
        ws.append([genre.genre_id, genre.title, genre.type, genre.rating, genre.description])

    # Сохраняем книгу и отправляем пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=genres.xlsx'
    wb.save(response)

    return response
@login_required
def report_genres(request):
    genres = Genre.objects.all()
    return render(request, 'genres/report_genres.html', {'genres': genres})


#Subscribers-----------------------------------------------------------------------------------------------------------------
@login_required
def display_subscribers(request):
    subscribers = Subscriber.objects.all()
    return render(request, 'subscribers/subscribers_list.html', {'subscribers': subscribers})
@login_required
def add_subscriber(request):
    if request.method == 'POST':
        form = SubscriberForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_subscribers')
    else:
        form = SubscriberForm()

    return render(request, 'subscribers/add_subscriber.html', {'form': form})
@login_required
def edit_subscriber(request, subscriber_id):
    subscriber = get_object_or_404(Subscriber, subscriber_id=subscriber_id)

    if request.method == 'POST':
        form = SubscriberForm(request.POST, instance=subscriber)
        if form.is_valid():
            form.save()
            return redirect('display_subscribers')
    else:
        form = SubscriberForm(instance=subscriber)

    return render(request, 'subscribers/edit_subscriber.html', {'form': form, 'subscriber': subscriber})
@login_required
def delete_subscriber(request, subscriber_id):
    subscriber = get_object_or_404(Subscriber, subscriber_id=subscriber_id)
    subscriber.delete()
    return redirect('display_subscribers')
@login_required
def export_subscribers_to_excel(request):
    subscribers = Subscriber.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['id', 'Имя', 'Фамилия', 'Адрес', 'Номер телефона'])

    for subscriber in subscribers:
        ws.append([subscriber.subscriber_id, subscriber.first_name, subscriber.last_name, subscriber.address, subscriber.phone_number])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=subscribers.xlsx'
    wb.save(response)

    return response
@login_required
def report_subscribers(request):
    subscribers = Subscriber.objects.all()
    return render(request, 'subscribers/report_subscribers.html', {'subscribers': subscribers})


#Book------------------------------------------------------------------------------------------------------------------------------------------


@login_required
def display_books(request):
    books = Book.objects.all()
    return render(request, 'books/books_list.html', {'books': books})
@login_required
def add_book(request):
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_books')
    else:
        form = BookForm()

    return render(request, 'books/add_book.html', {'form': form})
@login_required
def edit_book(request, book_id):
    book = get_object_or_404(Book, book_id=book_id)

    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
        if form.is_valid():
            form.save()
            return redirect('display_books')
    else:
        form = BookForm(instance=book)

    return render(request, 'books/edit_book.html', {'form': form, 'book': book})
@login_required
def delete_book(request, book_id):
    book = get_object_or_404(Book, book_id=book_id)
    book.delete()
    return redirect('display_books')
@login_required
def export_books_to_excel(request):
    books = Book.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['id', 'Название', 'Год выпуска', 'Автор', 'Жанр'])

    for book in books:
        ws.append([book.book_id, book.title, book.year, str(book.fk_author), str(book.fk_genre)])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=books.xlsx'
    wb.save(response)

    return response
@login_required
def report_books(request):
    books = Book.objects.all()
    return render(request, 'books/report_books.html', {'books': books})


#Book_Distribution-----------------------------------------------------------------------------------------------------------------------------------
@login_required
def display_book_distributions(request):
    book_distributions = BookDistribution.objects.all()
    return render(request, 'book_distributions/book_distributions_list.html', {'book_distributions': book_distributions})
@login_required
def add_book_distribution(request):
    if request.method == 'POST':
        form = BookDistributionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('display_book_distributions')
    else:
        form = BookDistributionForm()

    return render(request, 'book_distributions/add_book_distribution.html', {'form': form})
@login_required
def edit_book_distribution(request, book_distribution_id):
    book_distribution = get_object_or_404(BookDistribution, book_distribution_id=book_distribution_id)

    if request.method == 'POST':
        form = BookDistributionForm(request.POST, instance=book_distribution)
        if form.is_valid():
            form.save()
            return redirect('display_book_distributions')
    else:
        form = BookDistributionForm(instance=book_distribution)

    return render(request, 'book_distributions/edit_book_distribution.html', {'form': form, 'book_distribution': book_distribution})
@login_required
def delete_book_distribution(request, book_distribution_id):
    book_distribution = get_object_or_404(BookDistribution, book_distribution_id=book_distribution_id)
    book_distribution.delete()
    return redirect('display_book_distributions')
@login_required
def export_book_distributions_to_excel(request):
    book_distributions = BookDistribution.objects.all()

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Добавляем заголовки
    ws.append(['id', 'Дата выпуска', 'Дата возврата', 'Абонент', 'Сотрудник', 'Книга'])

    # Добавляем данные
    for book_distribution in book_distributions:
        ws.append([
            book_distribution.book_distribution_id,
            book_distribution.date_of_issue,
            book_distribution.return_date,
            str(book_distribution.fk_subscriber),
            str(book_distribution.fk_employee),
            str(book_distribution.fk_book)
        ])

    # Сохраняем книгу и отправляем пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=book_distributions.xlsx'
    wb.save(response)

    return response
@login_required
def report_book_distributions(request):
    book_distributions = BookDistribution.objects.all()
    return render(request, 'book_distributions/report_book_distributions.html', {'book_distributions': book_distributions})